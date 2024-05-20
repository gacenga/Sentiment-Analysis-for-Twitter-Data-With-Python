from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import os

def to_excel(png_paths, html_paths):
    """Saves visualizations into an excel file"""
    excel = 'sentiment_plots.xlsx'
    if os.path.exists(excel):
        workbook = load_workbook(excel)
    else:
        workbook = Workbook()
        workbook.active.title = "sentiment_plots"
    sheet = workbook.active

    paths = list(zip(png_paths, html_paths))
    for i, (png_path, html_path) in enumerate(zip(png_paths, html_paths)):
        #Insert images and links into the sheet
        img = Image(png_path)
        cell = f'A{1 + (i * 20)}'
        sheet.add_image(img, cell)
        sheet[f'A{2 + (i * 20)}'] = f'Interactive Plot {i + 1}'
        sheet[f'A{2 + (i * 20)}'].hyperlink = html_path
        sheet[f'A{2 + (i * 20)}'].style = "Hyperlink"

    #Save excel file
    workbook.save(excel)
